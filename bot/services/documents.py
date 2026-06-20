"""Document ingestion, chunking, FTS-backed retrieval, and Q&A.

Files sent by the user are stored under ``data/<user_id>/docs/``. Text is
extracted, chunked with overlap, indexed in ``document_chunks_fts``, and a
short LLM summary is produced on upload. Users ask questions by replying to
the summary message; the bot retrieves the most relevant chunks and answers
with an LLM prompt grounded in those chunks.
"""

import asyncio
import logging
import os
import shutil
from pathlib import Path
from typing import Any

from bot.ollama import OllamaChatMessage, generate_chat_completion
from bot.ollama.dto import OllamaErrorChunk
from bot.settings import (
    DOCUMENT_CHUNK_OVERLAP,
    DOCUMENT_CHUNK_SIZE,
    OLLAMA_MODEL,
)

logger = logging.getLogger(__name__)

db: Any = None  # injected at startup by bot.__init__

# Mapping of Telegram message_id -> document_id for reply-based Q&A.
# Kept in-process; if the bot restarts, users can still reference documents
# via /docs and reply chain by document id, but reply-to-summary is easiest.
_document_message_map: dict[int, int] = {}


SUPPORTED_TEXT_SUFFIXES = {
    ".txt",
    ".md",
    ".py",
    ".js",
    ".html",
    ".css",
    ".sql",
    ".log",
    ".xml",
    ".yaml",
    ".yml",
    ".json",
    ".csv",
    ".tsv",
}


def _extract_text_from_file(file_path: str, suffix: str) -> str:
    """Extract text from a local file path based on its extension."""
    suffix = suffix.lower()

    if suffix == ".pdf":
        try:
            import pypdf

            reader = pypdf.PdfReader(file_path)
            parts = []
            for i, page in enumerate(reader.pages):
                parts.append(page.extract_text() or "")
                if i >= 199:
                    parts.append("\n...[truncated at 200 pages]")
                    break
            return "\n".join(parts)
        except ImportError:
            return "[PDF: установите pypdf для извлечения текста]"
        except Exception as e:
            return f"[PDF extraction error: {e}]"

    if suffix == ".docx":
        try:
            import docx

            doc = docx.Document(file_path)
            return "\n".join(para.text for para in doc.paragraphs)
        except ImportError:
            return "[DOCX: установите python-docx для извлечения текста]"
        except Exception as e:
            return f"[DOCX extraction error: {e}]"

    if suffix in (".xlsx", ".xlsm"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                parts.append(f"# Лист: {sheet}")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if v is None else str(v) for v in row]
                    if any(cells):
                        parts.append("\t".join(cells))
                    row_count += 1
                    if row_count >= 1000:
                        parts.append("...[обрезано на 1000 строках]")
                        break
            wb.close()
            return "\n".join(parts)
        except ImportError:
            return "[XLSX: установите openpyxl для извлечения данных]"
        except Exception as e:
            return f"[XLSX extraction error: {e}]"

    if suffix in (".csv", ".tsv"):
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()

    if suffix == ".json":
        import json

        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        try:
            data = json.loads(content)
            return json.dumps(data, ensure_ascii=False, indent=2)
        except Exception:
            return content

    if suffix in SUPPORTED_TEXT_SUFFIXES:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()

    # Fallback: try plain text, otherwise report binary/unsupported.
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception:
        return f"[Unsupported or binary file type: {suffix}]"


def _detect_suffix(filename: str | None, mime_type: str | None) -> str:
    """Best-effort file extension detection."""
    if filename:
        suffix = Path(filename).suffix.lower()
        if suffix:
            return suffix
    mime_to_suffix = {
        "application/pdf": ".pdf",
        "text/plain": ".txt",
        "text/markdown": ".md",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
        "application/json": ".json",
        "text/csv": ".csv",
    }
    return mime_to_suffix.get((mime_type or "").lower(), ".bin")


def _chunk_text(
    text: str, chunk_size: int | None = None, overlap: int | None = None
) -> list[str]:
    """Split text into overlapping chunks on paragraph/word boundaries."""
    chunk_size = chunk_size or DOCUMENT_CHUNK_SIZE
    overlap = overlap if overlap is not None else DOCUMENT_CHUNK_OVERLAP
    if not text:
        return []

    text = text.replace("\r\n", "\n").strip()
    if len(text) <= chunk_size:
        return [text]

    # Split on paragraphs first; if a paragraph is longer than chunk_size,
    # split it on sentence boundaries, then on words as a last resort.
    paragraphs = [p.strip() for p in text.split("\n") if p.strip()]
    chunks: list[str] = []
    current: list[str] = []
    current_len = 0

    def _flush():
        nonlocal current, current_len
        if current:
            chunks.append("\n\n".join(current))
            current = []
            current_len = 0

    for para in paragraphs:
        para_len = len(para)
        if para_len > chunk_size:
            _flush()
            # Split oversized paragraph on sentence boundaries.
            sentences = []
            import re

            for sent in re.split(r"(?<=[.!?])\s+", para):
                sentences.append(sent)
            sub_current: list[str] = []
            sub_len = 0
            for sent in sentences:
                if sub_len + len(sent) + 1 > chunk_size and sub_current:
                    chunks.append(" ".join(sub_current))
                    # Carry over overlap from the end of the previous chunk.
                    overlap_text = " ".join(sub_current)[-overlap:]
                    sub_current = [overlap_text + sent] if overlap_text else [sent]
                    sub_len = len(sub_current[0])
                else:
                    sub_current.append(sent)
                    sub_len += len(sent) + 1
            if sub_current:
                chunks.append(" ".join(sub_current))
            continue

        if current_len + para_len + 2 > chunk_size and current:
            _flush()
            # Overlap: include tail of previous chunk.
            if chunks and overlap:
                tail = chunks[-1][-overlap:]
                if tail:
                    current.append(tail)
                    current_len += len(tail)

        current.append(para)
        current_len += para_len + 2

    _flush()
    return [c for c in chunks if c.strip()]


async def _ollama_simple_prompt(system: str, prompt: str, max_tokens: int = 600) -> str:
    """Run a short, non-streaming LLM call and return the text."""
    messages = [
        OllamaChatMessage(role="system", content=system),
        OllamaChatMessage(role="user", content=prompt),
    ]
    output = ""
    try:
        async with asyncio.timeout(60):
            async for is_done, chunk in generate_chat_completion(
                messages, OLLAMA_MODEL, temperature=0.3
            ):
                if is_done:
                    break
                if isinstance(chunk, OllamaErrorChunk):
                    logger.warning("[DOCS] LLM error: %s", chunk.error)
                    return ""
                output += chunk.message.content
    except asyncio.TimeoutError:
        logger.warning("[DOCS] LLM call timed out")
    except Exception as e:
        logger.warning("[DOCS] LLM call failed: %s", e)
    return output.strip()


async def summarize_text(text: str) -> str:
    """Produce a concise Russian summary of the document text."""
    if not text or not text.strip():
        return "Документ пуст."

    # Limit summary input so the LLM call stays fast.
    preview = text[:8000]
    prompt = (
        "Составь краткое содержание документа на русском языке (3–5 предложений). "
        "Опиши главную тему, ключевые выводы и полезные факты.\n\n"
        f"{preview}"
    )
    summary = await _ollama_simple_prompt(
        "Ты помощник, который делает краткие содержания документов.",
        prompt,
    )
    if not summary:
        # Fallback: first few paragraphs if LLM failed.
        paragraphs = [p for p in text.split("\n") if p.strip()][:3]
        summary = (
            "\n".join(paragraphs)
            if paragraphs
            else "[Не удалось сгенерировать краткое содержание]"
        )
    return summary


def _user_docs_dir(base_dir: str | Path, user_id: int) -> Path:
    path = Path(base_dir) / str(user_id) / "docs"
    path.mkdir(parents=True, exist_ok=True)
    return path


async def save_document(
    user_id: int,
    telegram_file_id: str | None,
    filename: str | None,
    mime_type: str | None,
    source_path: str,
    base_dir: str | Path,
) -> dict:
    """Persist a document: copy file, extract text, chunk, index, summarize."""
    if db is None:
        raise RuntimeError("Database not available")

    safe_name = Path(filename or "document").name
    suffix = _detect_suffix(filename, mime_type)
    docs_dir = _user_docs_dir(base_dir, user_id)
    local_filename = (
        f"{telegram_file_id or 'local'}_{safe_name}" if telegram_file_id else safe_name
    )
    local_path = str(docs_dir / local_filename)

    # Avoid overwriting by appending a counter.
    counter = 1
    original_local_path = local_path
    while Path(local_path).exists():
        stem = Path(original_local_path).stem
        ext = Path(original_local_path).suffix
        local_path = str(docs_dir / f"{stem}_{counter}{ext}")
        counter += 1

    shutil.copy2(source_path, local_path)

    text = _extract_text_from_file(local_path, suffix)
    chunks = _chunk_text(text)
    summary = await summarize_text(text)

    doc_id = db.add_document(
        user_id=user_id,
        telegram_file_id=telegram_file_id,
        local_path=local_path,
        filename=safe_name,
        mime_type=mime_type,
        extracted_text=text,
        summary=summary,
    )
    if chunks:
        db.add_document_chunks(doc_id, user_id, chunks)

    return {
        "id": doc_id,
        "user_id": user_id,
        "filename": safe_name,
        "local_path": local_path,
        "mime_type": mime_type,
        "summary": summary,
        "chunk_count": len(chunks),
        "text_length": len(text),
    }


def get_user_documents(user_id: int) -> list[dict]:
    if db is None:
        return []
    return db.get_documents(user_id)


def get_document(doc_id: int, user_id: int | None = None) -> dict | None:
    if db is None:
        return None
    return db.get_document(doc_id, user_id=user_id)


def delete_document(doc_id: int, user_id: int | None = None) -> bool:
    if db is None:
        return False
    doc = db.get_document(doc_id, user_id=user_id)
    if not doc:
        return False
    local_path = doc.get("local_path")
    if local_path and Path(local_path).exists():
        # Defensive guard: only delete files that belong to the user.
        if user_id is not None and db._is_path_inside_user_dir(local_path, user_id):
            try:
                os.unlink(local_path)
            except Exception as e:
                logger.warning("[DOCS] failed to remove file %s: %s", local_path, e)
        elif user_id is None:
            # Legacy/internal callers may pass no user_id (e.g. cascade cleanup
            # via Database.delete_user). Refuse to delete an unvalidated path.
            logger.warning(
                "[DOCS] delete_document called without user_id; skipping file removal"
            )
        else:
            logger.warning(
                "[DOCS] refusing to delete path outside user dir: %s", local_path
            )
    return db.delete_document(doc_id, user_id=user_id)


def search_document_chunks(user_id: int, query: str, limit: int = 5) -> list[dict]:
    if db is None or not query.strip():
        return []
    results = db.search_document_chunks(user_id, query, limit=limit)
    return results


def map_summary_message(message_id: int, doc_id: int) -> None:
    """Remember that a given Telegram message contains a document summary."""
    _document_message_map[message_id] = doc_id


def doc_id_for_message(message_id: int) -> int | None:
    """Resolve a document id from a summary message id (for reply-based Q&A)."""
    return _document_message_map.get(message_id)


async def answer_question(
    user_id: int, question: str, doc_id: int | None = None
) -> str:
    """Answer a question using retrieved chunks from the user's documents."""
    if db is None:
        return "⚠️ База данных недоступна."

    if doc_id is not None:
        doc = db.get_document(doc_id)
        if not doc or doc.get("user_id") != user_id:
            return "⚠️ Документ не найден или нет доступа."
        # Retrieve chunks from this document only.
        all_results = db.search_document_chunks(user_id, question, limit=10)
        results = [r for r in all_results if r.get("document_id") == doc_id][:5]
    else:
        results = db.search_document_chunks(user_id, question, limit=5)

    if not results:
        return "🔍 По твоему вопросу ничего не нашёл в загруженных документах."

    context = "\n\n---\n\n".join(r["chunk"] for r in results)
    prompt = (
        "Ответь на вопрос пользователя, опираясь только на приведённые фрагменты документа. "
        "Если в фрагментах нет ответа, так и скажи. Цитируй конкретные факты.\n\n"
        f"ФРАГМЕНТЫ:\n{context}\n\n"
        f"ВОПРОС: {question}\n\n"
        "ОТВЕТ:"
    )
    answer = await _ollama_simple_prompt(
        "Ты точный ассистент, который отвечает по документам.",
        prompt,
        max_tokens=1200,
    )
    if not answer:
        return "⚠️ Не удалось получить ответ от модели."
    return answer
