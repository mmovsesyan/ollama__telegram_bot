from aiogram import F, Router
from aiogram.filters.command import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import CallbackQuery, Message
from pydantic import BaseModel
from typing import Any
import asyncio
import json
import os
import tempfile
import time

from bot.bot import bot as aiogram_bot
from bot.keyboards.inline import answer_keyboard
from bot.keyboards.reply import command_keyboard, cancel_keyboard
from bot.ollama import OllamaChat, OllamaChatMessage, generate_chat_completion
from bot.ollama.api import get_installed_models, model_is_installed
from bot.ollama.dto import OllamaErrorChunk
from bot.security import is_allowed as _is_allowed
from bot.states import BotStates
from bot.settings import (
    COMPACTION_EVERY_N,
    MAX_CONTEXT_MESSAGES,
    MAX_CONTEXT_TOKENS,
    OLLAMA_MODEL,
    OLLAMA_MODEL_TEMPERATURE,
    START_USER_MESSAGE,
    SUMMARY_PROMPT,
    SYSTEM_MESSAGE,
    WHISPER_MODEL,
    WHISPER_DEVICE,
    WHISPER_COMPUTE_TYPE,
)

router = Router()

db = None  # injected in __init__

try:
    from faster_whisper import WhisperModel
    _WHISPER_AVAILABLE = True
except Exception:
    WhisperModel = None  # type: ignore[misc,assignment]
    _WHISPER_AVAILABLE = False

_whisper_model_instance: Any | None = None


def _get_whisper_model() -> Any:
    global _whisper_model_instance
    if _whisper_model_instance is None:
        if not _WHISPER_AVAILABLE or WhisperModel is None:
            raise RuntimeError("faster-whisper is not installed.")
        _whisper_model_instance = WhisperModel(
            WHISPER_MODEL,
            device=WHISPER_DEVICE,
            compute_type=WHISPER_COMPUTE_TYPE,
        )
    return _whisper_model_instance


def _escape_markdown(text: str) -> str:
    chars = r"_[]()~`>#+-=|{}.!"
    for ch in chars:
        text = text.replace(ch, "\\" + ch)
    return text


def wrap(s: str, w: int) -> list[str]:
    """Split a string into chunks of at most `w` characters, breaking on word
    boundaries when possible so emoji and grapheme clusters don't get cut.

    Falls back to hard chunking only if a single word exceeds `w` chars.
    """
    if not s:
        return []
    if len(s) <= w:
        return [s]
    out = []
    rem = s
    while len(rem) > w:
        # Look for the last whitespace within the window
        cut = rem.rfind(" ", 0, w)
        if cut <= 0:
            cut = rem.rfind("\n", 0, w)
        if cut <= 0:
            # No word boundary in range; hard-cut to avoid infinite loop
            cut = w
        out.append(rem[:cut].rstrip())
        rem = rem[cut:].lstrip()
    if rem:
        out.append(rem)
    return out


class UserChat(BaseModel):
    ollama_chat: OllamaChat
    selected_model: str = OLLAMA_MODEL
    linked_last_messages: int | None = None
    previous_prompt: str | None = None
    session_id: int | None = None
    last_active: float = 0


chats: dict[int, UserChat] = {}
_typing_last: dict[int, float] = {}
_request_last: dict[int, float] = {}
_generating: set[int] = set()


async def _cleanup_old_chats():
    now = time.time()
    stale = [uid for uid, chat in chats.items() if chat.last_active < now - 7200]
    for uid in stale:
        _delete_chat(uid)
        print(f"[CLEANUP] Removed idle session for user {uid}")


async def _safe_typing(user_id: int):
    now = time.time()
    if user_id in _typing_last and now - _typing_last[user_id] < 3:
        return
    _typing_last[user_id] = now
    try:
        await aiogram_bot.send_chat_action(chat_id=user_id, action="typing")
    except Exception as e:
        if "Flood control" in str(e) or "Too Many Requests" in str(e):
            pass
        else:
            print(f"[TYPING] Error: {e}")


def _estimate_tokens(text: str) -> int:
    return max(1, len(text) // 4)


BUTTON_MAP = {
    "💬 Чат": None,
    "🔍 Поиск": "/search",
    "🌤 Погода": "/weather",
    "⏰ Напомнить": "/remind",
    "📋 Задача": "/task",
    "📝 Заметка": "/note",
    "🧠 Память": "/memory",
    "📊 Отчёт": "/report",
    "❓ Помощь": "/help",
    "🗑 Очистить": "/clear",
}


async def generate(message: Message, user_id: int, text: str):
    now = time.time()
    if user_id in _request_last and now - _request_last[user_id] < 1:
        await message.answer("Слишком быстро. Подождите секунду.")
        return
    _request_last[user_id] = now

    if user_id in _generating:
        await message.answer("⏳ Подождите, я уже отвечаю...")
        return
    _generating.add(user_id)

    _create_chat(user_id)
    chat = chats[user_id]

    try:
        if chat.linked_last_messages:
            await aiogram_bot.edit_message_reply_markup(
                chat_id=user_id,
                message_id=chat.linked_last_messages,
                reply_markup=None,
            )
    except Exception:
        pass

    chat.linked_last_messages = None
    await aiogram_bot.send_chat_action(chat_id=user_id, action="typing")

    chat.last_active = time.time()
    prompt = text
    chat.previous_prompt = prompt
    chat.ollama_chat.messages.append(OllamaChatMessage(role="user", content=prompt))
    _trim_context(chat)

    print(f"[{user_id}]: {prompt}")

    if db and chat.session_id:
        db.save_message(user_id, chat.session_id, "user", prompt, chat.selected_model)

    msg = await message.answer("Думаю...")

    assistant_content = ""
    try:
        async with asyncio.timeout(300):
            async for is_done, chunk in generate_chat_completion(
                chat.ollama_chat.messages,
                chat.selected_model,
                temperature=OLLAMA_MODEL_TEMPERATURE,
            ):
                if is_done:
                    wrapped_response = wrap(assistant_content, 4096)
                    if not wrapped_response:
                        await msg.edit_text("(пустой ответ)", reply_markup=None)
                    else:
                        initial_content = wrapped_response.pop(0)
                        safe_text = _escape_markdown(initial_content)
                        try:
                            await msg.edit_text(
                                safe_text,
                                parse_mode="MarkdownV2",
                                reply_markup=None if wrapped_response else answer_keyboard,
                            )
                        except Exception as e:
                            print(f"Markdown error: {e}")
                            await msg.edit_text(
                                initial_content,
                                parse_mode=None,
                                reply_markup=None if wrapped_response else answer_keyboard,
                            )

                        for extra_text in wrapped_response:
                            extra_msg = await msg.answer(extra_text)
                            if wrapped_response.index(extra_text) == len(wrapped_response) - 1:
                                await extra_msg.edit_reply_markup(reply_markup=answer_keyboard)
                    print(f"[{user_id}]: Finished!")
                else:
                    if isinstance(chunk, OllamaErrorChunk):
                        await msg.edit_text(f"Ошибка Ollama: {chunk.error}")
                        break
                    assistant_content += chunk.message.content
                    if len(assistant_content) % 100 == 0:
                        await _safe_typing(user_id)
    except asyncio.TimeoutError:
        print(f"[ERROR] Generation timeout for user {user_id}")
        await msg.edit_text("⏳ Генерация заняла слишком много времени. Попробуйте ещё раз.")
        return
    except Exception as e:
        print(f"[ERROR] Generation failed: {e}")
        await msg.edit_text(f"Произошла ошибка при генерации ответа. Попробуйте ещё раз.\n({str(e)[:200]})")
        return
    finally:
        _generating.discard(user_id)

    chat.linked_last_messages = msg.message_id
    chat.ollama_chat.messages.append(
        OllamaChatMessage(role="assistant", content=assistant_content)
    )
    _trim_context(chat)

    if db and chat.session_id:
        db.save_message(user_id, chat.session_id, "assistant", assistant_content, chat.selected_model)

    # Fire-and-forget fact extraction so the user's reply isn't delayed.
    # Cheap LLM call (12s timeout, 0-3 facts per turn) populates the KB
    # so future "что я говорил про X" queries hit local cache.
    if db and assistant_content.strip():
        try:
            from bot.services.kb_extract import extract_facts_from_exchange
            asyncio.create_task(
                extract_facts_from_exchange(db, user_id, prompt, assistant_content)
            )
        except Exception:
            pass

    if db and chat.session_id:
        asyncio.create_task(_maybe_compact(user_id, chat))


def _trim_context(chat: UserChat) -> None:
    system_messages = [m for m in chat.ollama_chat.messages if m.role == "system"]
    other_messages = [m for m in chat.ollama_chat.messages if m.role != "system"]

    total_tokens = sum(_estimate_tokens(m.content) for m in system_messages)
    kept: list[OllamaChatMessage] = []
    for m in reversed(other_messages):
        tokens = _estimate_tokens(m.content)
        if total_tokens + tokens > MAX_CONTEXT_TOKENS and kept:
            break
        total_tokens += tokens
        kept.insert(0, m)

    chat.ollama_chat.messages = system_messages + kept


async def _maybe_compact(user_id: int, chat: UserChat):
    if not db or not chat.session_id:
        return

    non_system = [m for m in chat.ollama_chat.messages if m.role in ("user", "assistant")]
    total_count = len(non_system)

    if total_count < COMPACTION_EVERY_N:
        return

    latest_summary = db.get_latest_summary(chat.session_id)
    if latest_summary and latest_summary.get("message_count", 0) >= total_count:
        return

    print(f"[COMPACT] Triggered for user {user_id} at {total_count} messages")

    conversation_lines = []
    for m in non_system:
        role_label = "Пользователь" if m.role == "user" else "Ассистент"
        conversation_lines.append(f"{role_label}: {m.content}")
    conversation_text = "\n\n".join(conversation_lines)

    summary_prompt = (
        f"{SUMMARY_PROMPT}\n\n"
        f"ДИАЛОГ:\n{conversation_text}\n\n"
        f"ВЫЖИМКА:"
    )

    try:
        summary_messages = [
            OllamaChatMessage(role="system", content=SYSTEM_MESSAGE),
            OllamaChatMessage(role="user", content=summary_prompt),
        ]
        summary_content = ""
        async for is_done, chunk in generate_chat_completion(
            summary_messages,
            chat.selected_model,
            temperature=0.3,
        ):
            if is_done:
                break
            else:
                if isinstance(chunk, OllamaErrorChunk):
                    print(f"[COMPACT] Summary error: {chunk.error}")
                    return
                summary_content += chunk.message.content

        if not summary_content.strip():
            print("[COMPACT] Empty summary, skipping")
            return

        db.add_summary(chat.session_id, total_count, summary_content)
        print(f"[COMPACT] Saved summary for session {chat.session_id} at {total_count} messages")

        memory_prompt = (
            "Проанализируй диалог и извлеки ВАЖНЫЕ факты о пользователе.\n"
            "Для каждого факта укажи категорию: fact (факт), preference (предпочтение), note (заметка).\n"
            "Ответь ТОЛЬКО в формате (один факт на строку):\n"
            "[category] content\n"
            "Если нет важных фактов — напиши \"НЕТ\".\n\n"
            f"ДИАЛОГ:\n{conversation_text}\n\n"
            "ФАКТЫ:"
        )
        try:
            memory_messages = [
                OllamaChatMessage(role="system", content=SYSTEM_MESSAGE),
                OllamaChatMessage(role="user", content=memory_prompt),
            ]
            memory_raw = ""
            async for is_done, chunk in generate_chat_completion(
                memory_messages,
                chat.selected_model,
                temperature=0.2,
            ):
                if is_done:
                    break
                else:
                    if isinstance(chunk, OllamaErrorChunk):
                        break
                    memory_raw += chunk.message.content

            if memory_raw.strip() and memory_raw.strip().upper() != "НЕТ":
                for line in memory_raw.strip().split("\n"):
                    line = line.strip()
                    if not line or line.upper() == "НЕТ":
                        continue
                    if line.startswith("[") and "]" in line:
                        close_idx = line.index("]")
                        category = line[1:close_idx].lower().strip()
                        content = line[close_idx+1:].strip()
                        if category in ("fact", "preference", "note") and content:
                            existing = db.get_memories(user_id, category)
                            dup = any(m.get('content','').lower() == content.lower() for m in existing)
                            if not dup:
                                db.add_memory(user_id, category, content)
                                print(f"[MEMORY] Auto-saved: [{category}] {content}")
        except Exception as mem_err:
            print(f"[MEMORY] Extraction failed: {mem_err}")

        # Background pass: any long memories (>500 chars) without a summary
        # get one. Bounded to 5/run so a backlog doesn't burn tokens.
        try:
            from bot.services.kb_extract import compress_pending_memories
            asyncio.create_task(compress_pending_memories(db, user_id, limit=5))
        except Exception:
            pass

        base_system_msgs = [m for m in chat.ollama_chat.messages if m.role == "system"][:1]
        summary_msg = OllamaChatMessage(
            role="system",
            content=f"[Контекст предыдущего диалога]: {summary_content}"
        )
        last_pairs = non_system[-4:]

        chat.ollama_chat.messages = base_system_msgs + [summary_msg] + [
            OllamaChatMessage(role=m.role, content=m.content) for m in last_pairs
        ]
        print(f"[COMPACT] Context rebuilt: {len(chat.ollama_chat.messages)} messages")
    except Exception as e:
        print(f"[COMPACT] Failed: {e}")


def _delete_chat(user_id: int) -> None:
    if user_id not in chats:
        return
    del chats[user_id]


def _build_system_content(user_id: int) -> str:
    """Build the dynamic system prompt: base message + notes + memories.

    Kept as a pure function so it can be reused when refreshing a live chat
    after a new note/memory is saved."""
    system_content = SYSTEM_MESSAGE
    if db is None:
        return system_content

    notes = db.get_notes(user_id)
    if notes:
        system_content += f"\n\nКонтекст о пользователе:\n{notes}"

    memories = db.get_memories(user_id)
    if memories:
        memory_lines = []
        for m in memories:
            cat = m.get('category', 'fact')
            content = m.get('content', '')
            summary = m.get('summary')
            display = summary if summary else content
            memory_lines.append(f"- [{cat}] {display}")
        system_content += "\n\nВажные факты и предпочтения:\n" + "\n".join(memory_lines)

    return system_content


def _find_summary_message(messages: list[OllamaChatMessage]) -> OllamaChatMessage | None:
    """Return the first non-base system message that carries a previous-dialog
    summary marker, or None if absent."""
    for m in messages:
        if m.role == "system" and m.content.startswith("[Контекст предыдущего диалога]:"):
            return m
    return None


def refresh_system_prompt(user_id: int) -> bool:
    """Reload notes/memories into an active chat's system prompt in-place.

    Called after any action that updates persistent user context (notes,
    memories, auto-extracted facts) so the next LLM call already knows the
    new information without requiring /clear or session timeout.

    Preserves the base system message and any previous-dialog summary."""
    chat = chats.get(user_id)
    if chat is None or db is None:
        return False

    base_system = None
    summary_msg = _find_summary_message(chat.ollama_chat.messages)
    other_messages = [m for m in chat.ollama_chat.messages if m.role != "system"]

    # The first system message is always the base prompt built by _build_system_content.
    if chat.ollama_chat.messages and chat.ollama_chat.messages[0].role == "system":
        base_system = chat.ollama_chat.messages[0]

    new_system_content = _build_system_content(user_id)
    if base_system is None:
        chat.ollama_chat.messages.insert(
            0, OllamaChatMessage(role="system", content=new_system_content)
        )
    else:
        base_system.content = new_system_content

    # Ensure summary sits right after the base system prompt.
    if summary_msg is not None:
        # Re-insert if it got dropped during rebuild.
        if summary_msg not in chat.ollama_chat.messages:
            chat.ollama_chat.messages.insert(1, summary_msg)
        else:
            # Move to position 1 if not already there.
            idx = chat.ollama_chat.messages.index(summary_msg)
            if idx != 1:
                chat.ollama_chat.messages.pop(idx)
                chat.ollama_chat.messages.insert(1, summary_msg)

    # Trim in case refreshing pushed us over budget.
    _trim_context(chat)
    return True


def _create_chat(user_id: int) -> bool:
    if user_id in chats:
        return False

    session_id = None
    if db:
        session_id = db.get_or_create_active_session(user_id, OLLAMA_MODEL)

    history = []
    if db:
        history = db.get_session_messages(user_id, limit=MAX_CONTEXT_MESSAGES)

    chats[user_id] = UserChat(
        selected_model=OLLAMA_MODEL,
        ollama_chat=OllamaChat(messages=[]),
        session_id=session_id,
    )

    system_content = _build_system_content(user_id)
    if system_content:
        chats[user_id].ollama_chat.messages.append(
            OllamaChatMessage(role="system", content=system_content)
        )

    if db and session_id:
        latest_summary = db.get_latest_summary(session_id)
        if latest_summary and latest_summary.get("summary"):
            chats[user_id].ollama_chat.messages.append(
                OllamaChatMessage(
                    role="system",
                    content=f"[Контекст предыдущего диалога]: {latest_summary['summary']}"
                )
            )

    for h in history:
        chats[user_id].ollama_chat.messages.append(
            OllamaChatMessage(role=h["role"], content=h["content"])
        )

    if START_USER_MESSAGE:
        chats[user_id].ollama_chat.messages.append(
            OllamaChatMessage(role="user", content=START_USER_MESSAGE)
        )
    return True


@router.message(Command("models"))
@router.message(lambda m: m.text and m.text == "/models")
@router.message(F.text == "🤖 Модели")
async def cmd_models(message: Message):
    if message.from_user is None:
        return
    if not _is_allowed(message.from_user.id):
        print(f"[BLOCKED] Unauthorized user {message.from_user.id}")
        return
    models = await get_installed_models()
    model_list = "\n".join([f"- {m.name}" for m in models]) or "Нет моделей"
    await message.answer(f"Доступные модели:\n{model_list}", reply_markup=command_keyboard)


@router.message(Command("help"))
@router.message(lambda m: m.text and m.text == "/help")
@router.message(F.text == "❓ Помощь")
async def cmd_help(message: Message):
    if message.from_user is None:
        return
    if not _is_allowed(message.from_user.id):
        print(f"[BLOCKED] Unauthorized user {message.from_user.id}")
        return
    await message.answer(
        "🤖 Вот что я умею:\n\n"
        "🌤 *Погода*\n"
        "• «погода в Москве»\n\n"
        "⏰ *Напоминания*\n"
        "• «напомни через 5 минут позвонить»\n"
        "• «завтра в 9:00 проверить отчёт»\n"
        "• «каждое утро в 9 покажи новости»\n\n"
        "📋 *Задачи (AI выполнит сам)*\n"
        "• «задача каждый день в 7:00 погода в Москве»\n"
        "• «задача через час поищи новости Tesla»\n\n"
        "📝 *Заметки*\n"
        "• «заметка: купить акции TSLA»\n\n"
        "🧠 *Память*\n"
        "• «запомни, я люблю краткие ответы»\n"
        "• «факт: я работаю над проектом X»\n\n"
        "🔍 *Поиск и новости*\n"
        "• «поищи последние новости Tesla»\n"
        "• «новости»\n\n"
        "💬 *AI-чат*\n"
        "• просто напиши вопрос — бот ответит через Ollama\n\n"
        "📋 *Команды:*\n"
        "/start — меню\n"
        "/remind — напоминание\n"
        "/task — задача\n"
        "/note — заметка\n"
        "/memory — память\n"
        "/models — модели\n"
        "/model — сменить модель\n"
        "/clear — очистить историю\n"
        "/monitors — мониторы",
        reply_markup=command_keyboard,
        parse_mode="Markdown",
    )


@router.message(Command("model"))
@router.message(lambda m: m.text and (m.text == "/model" or m.text.startswith("/model ")))
async def cmd_model(message: Message, state: FSMContext):
    if message.from_user is None:
        return
    user_id = message.from_user.id
    if not _is_allowed(user_id):
        print(f"[BLOCKED] Unauthorized user {user_id}")
        return
    parts = message.text.split(maxsplit=1)
    if len(parts) < 2:
        await message.answer(
            "Введи название модели:\n"
            "Пример: llama2:13b-chat",
            reply_markup=cancel_keyboard,
        )
        await state.set_state(BotStates.waiting_model)
        return
    model_to_set = parts[1].strip()
    if not await model_is_installed(model_to_set):
        await message.answer(f"Модель {model_to_set} не найдена!", reply_markup=command_keyboard)
        return
    _create_chat(user_id)
    chats[user_id].selected_model = model_to_set
    await message.answer(f"✅ Модель изменена на {model_to_set}", reply_markup=command_keyboard)


@router.message(BotStates.waiting_model)
async def process_model_state(message: Message, state: FSMContext):
    if message.from_user is None:
        await state.clear()
        return
    user_id = message.from_user.id
    if not _is_allowed(user_id):
        await state.clear()
        return
    if message.text is None:
        await message.answer("Ожидался текст.", reply_markup=cancel_keyboard)
        await state.clear()
        return
    text = message.text or ""
    if text == "❌ Отмена":
        await state.clear()
        await message.answer("Действие отменено.", reply_markup=command_keyboard)
        return
    if text.startswith("/") or text in BUTTON_MAP:
        await state.clear()
        await message.answer("Текущее действие отменено.", reply_markup=command_keyboard)
        return
    model_to_set = text.strip()
    if not await model_is_installed(model_to_set):
        await message.answer(f"Модель {model_to_set} не найдена!", reply_markup=command_keyboard)
        await state.clear()
        return
    _create_chat(user_id)
    chats[user_id].selected_model = model_to_set
    await message.answer(f"✅ Модель изменена на {model_to_set}", reply_markup=command_keyboard)
    await state.clear()


@router.message(Command("clear"))
@router.message(lambda m: m.text and m.text == "/clear")
@router.message(F.text == "🗑 Очистить")
async def cmd_clear(message: Message):
    if message.from_user is None:
        return
    user_id = message.from_user.id
    if not _is_allowed(user_id):
        print(f"[BLOCKED] Unauthorized user {user_id}")
        return
    if user_id in chats and chats[user_id].session_id and db:
        db.close_session(chats[user_id].session_id, "User cleared chat")
    _delete_chat(user_id)
    await message.answer("✅ История очищена.", reply_markup=command_keyboard)


@router.message(F.text == "💬 Чат")
async def btn_chat(message: Message):
    if message.from_user is None:
        return
    if not _is_allowed(message.from_user.id):
        print(f"[BLOCKED] Unauthorized user {message.from_user.id}")
        return
    await message.answer(
        "💬 Просто напиши или скажи голосом, что нужно.\n\n"
        "Например:\n"
        "• «погода в Москве»\n"
        "• «напомни через 5 минут позвонить»\n"
        "• «поищи последние новости Tesla»",
        reply_markup=command_keyboard,
    )


@router.callback_query(F.data == "like")
async def like(callback: CallbackQuery):
    if not callback.from_user:
        return print("[ERROR]: Invalid message")

    user_id = callback.from_user.id
    if user_id not in chats:
        return

    chat = chats[user_id]
    if chat.linked_last_messages:
        try:
            await aiogram_bot.edit_message_reply_markup(
                chat_id=user_id,
                message_id=chat.linked_last_messages,
                reply_markup=None,
            )
        except Exception:
            pass
    chat.linked_last_messages = None
    await callback.answer("👍")


@router.callback_query(F.data == "dislike")
async def dislike(callback: CallbackQuery):
    if not callback.from_user:
        return print("[ERROR]: Invalid message")

    user_id = callback.from_user.id
    if user_id not in chats:
        return

    chat = chats[user_id]
    if not chat.linked_last_messages:
        return
    try:
        await aiogram_bot.delete_message(
            user_id,
            message_id=chat.linked_last_messages,
        )
    except Exception:
        pass
    if not chat.previous_prompt:
        return
    if not isinstance(callback.message, Message):
        raise Exception

    await generate(callback.message, user_id, chat.previous_prompt)
    await callback.answer("Перегенерация...")


async def _download_document(document):
    file = await aiogram_bot.get_file(document.file_id)
    suffix = os.path.splitext(document.file_name or "")[1] or ".bin"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp_path = tmp.name
    await aiogram_bot.download_file(file.file_path, tmp_path)
    return tmp_path, document.file_name or "file", suffix.lower()


def _extract_text_from_file(file_path: str, suffix: str) -> str:
    if suffix == ".pdf":
        try:
            import pypdf
            reader = pypdf.PdfReader(file_path)
            parts = []
            for i, page in enumerate(reader.pages):
                parts.append(page.extract_text() or "")
                if i >= 30:
                    parts.append("\n...[truncated at 30 pages]")
                    break
            return "\n".join(parts)
        except ImportError:
            return "[PDF: установите pypdf для извлечения текста]"
        except Exception as e:
            return f"[PDF extraction error: {e}]"
    elif suffix == ".json":
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
        try:
            data = json.loads(content)
            return json.dumps(data, ensure_ascii=False, indent=2)
        except Exception:
            return content
    elif suffix in (".csv", ".tsv"):
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    elif suffix in (".txt", ".md", ".py", ".js", ".html", ".css", ".sql", ".log", ".xml", ".yaml", ".yml"):
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    elif suffix == ".docx":
        try:
            import docx
            doc = docx.Document(file_path)
            return "\n".join(para.text for para in doc.paragraphs)
        except ImportError:
            return "[DOCX: установите python-docx для извлечения текста]"
        except Exception as e:
            return f"[DOCX extraction error: {e}]"
    elif suffix in (".xlsx", ".xlsm"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            parts = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                parts.append(f"# Лист: {sheet}")
                row_count = 0
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if v is None else str(v) for v in row]
                    if any(cells):
                        parts.append("\t".join(cells))
                    row_count += 1
                    if row_count >= 500:
                        parts.append("...[обрезано на 500 строках]")
                        break
            wb.close()
            return "\n".join(parts)
        except ImportError:
            return "[XLSX: установите openpyxl для извлечения данных]"
        except Exception as e:
            return f"[XLSX extraction error: {e}]"
    else:
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                return f.read()
        except Exception:
            return f"[Unsupported or binary file type: {suffix}]"


@router.message(F.document)
async def handle_document(message: Message, state: FSMContext):
    if message.from_user is None:
        return
    user_id = message.from_user.id
    if not _is_allowed(user_id):
        return
    document = message.document
    if document is None:
        return

    await message.answer(f"📄 Получен файл: {document.file_name or 'unknown'}\nЗагружаю и извлекаю текст...")

    tmp_path = None
    try:
        tmp_path, fname, suffix = await _download_document(document)
        text = _extract_text_from_file(tmp_path, suffix)
    except Exception as e:
        await message.answer(f"❌ Ошибка обработки файла: {str(e)[:200]}", reply_markup=command_keyboard)
        return
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not text or not text.strip():
        await message.answer("❌ Не удалось извлечь текст из файла.", reply_markup=command_keyboard)
        return

    max_chars = 12000
    original_len = len(text)
    if len(text) > max_chars:
        text = text[:max_chars] + f"\n\n...[обрезано с {original_len} символов]"

    prompt = f"[Документ: {fname}]\n\n{text}\n\nПроанализируй содержимое и дай краткий обзор."
    await generate(message, user_id, prompt)


async def _download_tg_file(file_id: str, suffix: str) -> tuple[str, str]:
    tg_file = await aiogram_bot.get_file(file_id)
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        tmp_path = tmp.name
    await aiogram_bot.download_file(tg_file.file_path, tmp_path)
    return tmp_path, tg_file.file_path or ""


async def _transcribe_audio(file_path: str) -> str:
    model = _get_whisper_model()
    segments, _info = model.transcribe(file_path, beam_size=5, vad_filter=True)
    return " ".join(segment.text for segment in segments).strip()


async def _handle_voice_or_audio(message: Message, state: FSMContext, file_id: str, suffix: str, label: str):
    if message.from_user is None:
        return
    user_id = message.from_user.id
    if not _is_allowed(user_id):
        return

    if not _WHISPER_AVAILABLE:
        await message.answer(
            "🎤 Распознавание голоса недоступно. Установи faster-whisper:\n"
            "poetry install --no-dev",
            reply_markup=command_keyboard,
        )
        return

    status_msg = await message.answer("🎤 Слушаю и распознаю речь...")
    tmp_path = None
    try:
        tmp_path, _ = await _download_tg_file(file_id, suffix)
        text = await _transcribe_audio(tmp_path)
    except Exception as e:
        err_text = str(e)
        if "ffmpeg" in err_text.lower() or "command not found" in err_text.lower():
            await status_msg.edit_text(
                "❌ Для распознавания голоса нужен ffmpeg.\n"
                "macOS: brew install ffmpeg\n"
                "Linux: sudo apt install ffmpeg"
            )
        else:
            await status_msg.edit_text(f"❌ Ошибка распознавания {label}: {err_text[:200]}")
        print(f"[VOICE] Transcription failed: {e}")
        return
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)

    if not text:
        await status_msg.edit_text("❌ Не удалось распознать речь. Попробуй ещё раз.")
        return

    await status_msg.edit_text(f"🎤 Распознано: {text}")
    # Treat the transcribed text as if the user typed it. Routing into the
    # smart-handler pipeline keeps voice and text on the same path.
    from bot.handlers.smart import smart_message_handler
    from copy import copy
    voice_msg = copy(message)
    voice_msg.text = text
    await smart_message_handler(voice_msg, state=state)


@router.message(F.voice)
async def handle_voice(message: Message, state: FSMContext):
    voice = message.voice
    if voice is None:
        return
    await _handle_voice_or_audio(message, state, voice.file_id, ".ogg", "голосового сообщения")


@router.message(F.audio)
async def handle_audio(message: Message, state: FSMContext):
    audio = message.audio
    if audio is None:
        return
    suffix = ".mp3" if (audio.file_name or "").lower().endswith(".mp3") else ".audio"
    await _handle_voice_or_audio(message, state, audio.file_id, suffix, "аудио")



@router.errors()
async def global_error_handler(event):
    update = event.update
    exception = event.exception
    print(f"[GLOBAL ERROR] {exception}")
    if update.message:
        try:
            from bot.keyboards.reply import command_keyboard
            await update.message.answer(
                "⚠️ Произошла ошибка. Попробуй ещё раз или используй /help.",
                reply_markup=command_keyboard,
            )
        except Exception as e:
            print(f"[ERROR HANDLER FAIL] {e}")
